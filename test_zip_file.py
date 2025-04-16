import csv
import os
import zipfile
from pypdf import PdfReader
from openpyxl import load_workbook

CURRENT_FILE = os.path.abspath(__file__) # получаем абсолютный путь к текущему файлу
CURRENT_DIRECTORY = os.path.dirname(CURRENT_FILE) # получаем абсолютный путь к текущей директории где находится файл с которым работаем
RESOURCES_DIR = os.path.join(CURRENT_DIRECTORY, 'resources') # делаем склейку пути к текущей директории и папки resources
FILES_DIR = os.path.join(CURRENT_DIRECTORY, 'files') # делаем склейку пути к текущей директории и папки files


def test_csv(create_archive):
    archive_path = os.path.join(RESOURCES_DIR, 'project_archive.zip')
    with zipfile.ZipFile(archive_path) as zip_file:
        with zip_file.open('Mock data.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            header_row = csvreader[0]
            second_row = csvreader[1]
            third_row = csvreader[2]
            fourth_row = csvreader[3]
            fifth_row = csvreader[4]

            assert header_row[2] == 'last_name'
            assert second_row[2] == 'Hrus'
            assert third_row[2] == 'Gaskill'
            assert fourth_row[2] == 'Garthside'
            assert fifth_row[2] == 'Castanone'

def test_xlsx(create_archive):
    archive_path = os.path.join(RESOURCES_DIR, 'project_archive.zip')
    with zipfile.ZipFile(archive_path) as zip_file:
        with zip_file.open('Programming Language.xlsx') as excel_file:
            sheet = load_workbook(excel_file).active

            assert sheet.cell(row=1, column=2).value == 'Язык программирования'
            assert sheet.cell(row=2, column=2).value == 'Python'
            assert sheet.cell(row=8, column=4).value == 74
            assert sheet.cell(row=10, column=2).value == 'Kotlin'

def test_pdf(create_archive):
    archive_path = os.path.join(RESOURCES_DIR, 'project_archive.zip')
    with zipfile.ZipFile(archive_path) as zip_file:
        with zip_file.open('Python or Go.pdf') as pdf_file:
            pdf_reader = PdfReader(pdf_file)

            full_text = ""
            for page in pdf_reader.pages:
                full_text += page.extract_text()

            assert 'Python: Simple and Powerful' in full_text
            assert 'Почему Python отлично подходит для веб-скрейпинга' in full_text
